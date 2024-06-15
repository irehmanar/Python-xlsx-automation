import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_company = {}
inventory_per_company = {}

product_list.cell(1,5).value ="Changed data"
for row in range(2,product_list.max_row+1):
    product_list.cell(row,5).value = product_list.cell(row,2).value * product_list.cell(row,3).value
    inventory = product_list.cell(row,2).value
    company_name = product_list.cell(row,4).value
    if company_name in product_per_company:
        product_per_company[company_name] = product_per_company[company_name] + 1
    else:
        product_per_company[company_name] = 1


    if company_name in inventory_per_company:
        inventory_per_company[company_name] = inventory_per_company[company_name] + inventory
    else:
        inventory_per_company[company_name] = inventory

    
print(product_per_company)
print(inventory_per_company)
inv_file.save("changed.xlsx")
